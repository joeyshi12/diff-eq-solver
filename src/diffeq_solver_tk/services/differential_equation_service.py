from abc import abstractmethod, ABC
from typing import TypeVar, Generic, Optional

import numpy as np
from matplotlib import cm
from matplotlib.animation import FuncAnimation
from matplotlib.axes import Axes
from matplotlib.figure import Figure
from openpyxl import Workbook

from diffeq_solver_tk.differential_equation_metadata import OrdinaryDifferentialEquationMetadata, BoundedEquationMetadata

T = TypeVar('T')


class DifferentialEquationService(Generic[T], ABC):
    metadata: Optional[T] = None
    solution: Optional[np.ndarray] = None
    main_figure: Figure

    def __init__(self, main_figure: Figure):
        self.main_figure = main_figure

    def compute_and_update_solution(self, metadata: T):
        self.validate_metadata(metadata)
        self.metadata = metadata
        self.solution = self.compute_solution(metadata)
        self.render_current_solution()

    def clear_solution(self):
        self.metadata = None
        self.solution = None
        self.clear_figure()

    def clear_figure(self):
        self.main_figure.clf()

    @abstractmethod
    def compute_solution(self, metadata: T) -> np.ndarray:
        raise NotImplementedError

    @abstractmethod
    def export_solution(self, table_path: str):
        raise NotImplementedError

    @abstractmethod
    def render_current_solution(self):
        raise NotImplementedError

    @abstractmethod
    def validate_metadata(self, metadata: T):
        raise NotImplementedError


class OrdinaryDifferentialEquationService(DifferentialEquationService[OrdinaryDifferentialEquationMetadata], ABC):
    def __init__(self, main_figure: Figure):
        super().__init__(main_figure)

    def render_current_solution(self):
        domain = np.linspace(0, self.metadata.time, self.metadata.samples)
        ax = self.main_figure.add_subplot()
        ax.plot(domain, self.solution)
        ax.set_xlabel("t [time]")
        ax.set_ylabel("x(t)")
        ax.plot()
        ax.set_title("Solution: x(t)")

    def export_solution(self, table_path: str):
        domain = np.linspace(0, self.metadata.time, self.metadata.samples)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.cell(1, 1, 't')
        worksheet.cell(1, 2, 'x')
        row = 2
        for t, x in zip(domain, self.solution):
            worksheet.cell(row, 1, t)
            worksheet.cell(row, 2, x)
            row += 1
        workbook.save(table_path)


class BoundedEquationService(DifferentialEquationService[BoundedEquationMetadata], ABC):
    __animation: Optional[FuncAnimation] = None
    __is_animation_playing: bool = False

    def __init__(self, main_figure: Figure):
        super().__init__(main_figure)

    def render_current_solution(self):
        self.clear_figure()
        self.clear_animation()
        K, N = self.solution.shape
        time_domain = np.linspace(0, self.metadata.time, K)
        space_domain = np.linspace(0, self.metadata.length, N)
        x, t = np.meshgrid(space_domain, time_domain)
        ax = self.main_figure.add_subplot(projection='3d')
        ax.set_xlabel('x [length]')
        ax.set_ylabel('t [time]')
        ax.set_zlabel('u(x, t)')
        surf = ax.plot_surface(x, t, self.solution, cmap=cm.coolwarm, linewidth=0, antialiased=False)
        self.main_figure.colorbar(surf, shrink=0.5, aspect=5)

    def generate_animation(self):
        self.clear_figure()
        K, N = self.solution.shape
        ax = self.main_figure.add_subplot()
        Axes.set_xlim(ax, left=0, right=self.metadata.length)
        Axes.set_ylim(ax, bottom=np.min(self.solution), top=np.max(self.solution))
        ax.set_xlabel("x [length]")
        ax.set_ylabel("u(x, t)")
        ax.set_title("Solution: u(x, t)")
        space_domain = np.linspace(0, self.metadata.length, self.metadata.samples)
        line, = ax.plot(space_domain, self.solution[0], "-o", markersize=4)
        time_text = ax.text(0.82, 0.92, '', transform=ax.transAxes)
        dt = self.metadata.time / K

        def update_plot(k: int):
            line.set_ydata(self.solution[k])
            time_text.set_text("t = %.3f" % (k * dt))
            return line, time_text

        self.__animation = FuncAnimation(self.main_figure, update_plot, frames=K, blit=True, interval=20)

    def toggle_animation(self):
        if not self.__animation:
            self.generate_animation()
            self.__is_animation_playing = True
        elif self.is_animation_playing():
            self.__animation.pause()
            self.__is_animation_playing = False
        else:
            self.__animation.resume()
            self.__is_animation_playing = True

    def is_animation_playing(self):
        return self.__is_animation_playing

    def export_solution(self, table_path: str):
        K, N = self.solution.shape
        time_domain = np.linspace(0, self.metadata.time, K)
        space_domain = np.linspace(0, self.metadata.length, N)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.cell(1, 2, "x →")
        for i in range(N):
            worksheet.cell(1, 3 + i, space_domain[i])
        worksheet.write(2, 1, "t ↓")
        for k in range(K):
            worksheet.write(3 + k, 1, time_domain[k])
        for k in range(K):
            for i in range(N):
                worksheet.write(k + 3, i + 3, self.solution[k, i])
        workbook.save(table_path)

    def clear_animation(self):
        if self.__animation:
            self.__animation.pause()
            self.__animation = None
            self.__is_animation_playing = False
            self.main_figure.clf()

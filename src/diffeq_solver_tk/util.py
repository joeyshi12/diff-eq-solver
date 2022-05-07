import numpy as np
from openpyxl import Workbook


def export_ode_solution(solution, metadata, table_path: str):
    domain = np.linspace(0, metadata.time, metadata.samples)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.cell(1, 1, 't')
    worksheet.cell(1, 2, 'x')
    row = 2
    for t, x in zip(domain, solution):
        worksheet.cell(row, 1, t)
        worksheet.cell(row, 2, x)
        row += 1
    workbook.save(table_path)


def export_bounded_equation_solution(solution, metadata, table_path: str):
    K, N = solution.shape
    time_domain = np.linspace(0, metadata.time, K)
    space_domain = np.linspace(0, metadata.length, N)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.cell(1, 2, "x →")
    for i in range(N):
        worksheet.cell(1, 3 + i, space_domain[i])
    worksheet.write(2, 1, "t ↓")
    for k in range(K):
        worksheet.write(3 + k, 1, time_domain[k])
    for k in range(K):
        for i in range(N):
            worksheet.write(k + 3, i + 3, solution[k, i])
    workbook.save(table_path)
